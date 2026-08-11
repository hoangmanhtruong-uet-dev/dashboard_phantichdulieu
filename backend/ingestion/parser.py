import csv
import re
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from backend.config import settings


SUPPORTED_TYPES = {"STRING", "NUMBER", "BOOLEAN", "DATE_TIME", "CURRENCY"}
CURRENCY_PATTERN = re.compile(
    r"^\s*(?:[$€£¥₫]|VND|USD|EUR|GBP)?\s*[-+]?\d[\d,.\s]*\s*(?:[$€£¥₫]|VND|USD|EUR|GBP)?\s*$",
    re.IGNORECASE,
)


class IngestionError(Exception):
    def __init__(self, code: str, message: str, status_code: int = 422):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    values: dict[str, Any]
    formula_fields: frozenset[str]


def _safe_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _validate_headers(values: list[Any]) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in values]
    if not headers or not any(headers):
        raise IngestionError("MISSING_HEADER", "The file does not contain a header row")
    if any(not header for header in headers):
        raise IngestionError(
            "EMPTY_COLUMN_NAME", "Every column must have a non-empty name"
        )
    if len(headers) > settings.max_import_columns:
        raise IngestionError(
            "TOO_MANY_COLUMNS",
            f"Files may contain at most {settings.max_import_columns} columns",
            413,
        )
    duplicates = [name for name, count in Counter(headers).items() if count > 1]
    if duplicates:
        raise IngestionError(
            "DUPLICATE_COLUMN", f"Duplicate column names: {', '.join(duplicates[:5])}"
        )
    return headers


def _validate_cell(value: Any) -> None:
    if isinstance(value, str) and len(value) > settings.max_cell_chars:
        raise IngestionError(
            "CELL_TOO_LARGE",
            f"A cell exceeds the {settings.max_cell_chars} character limit",
            413,
        )


def _csv_dialect(path: Path) -> type[csv.Dialect]:
    try:
        with path.open(
            "r", encoding="utf-8-sig", errors="strict", newline=""
        ) as handle:
            sample = handle.read(8192)
    except UnicodeDecodeError as exc:
        raise IngestionError(
            "INVALID_ENCODING", "CSV files must use UTF-8 encoding"
        ) from exc
    if not sample.strip():
        raise IngestionError("EMPTY_FILE", "The uploaded file is empty")
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _iter_csv(path: Path) -> Iterator[ParsedRow]:
    dialect = _csv_dialect(path)
    try:
        with path.open(
            "r", encoding="utf-8-sig", errors="strict", newline=""
        ) as handle:
            reader = csv.reader(handle, dialect=dialect, strict=True)
            try:
                headers = _validate_headers(next(reader))
            except StopIteration as exc:
                raise IngestionError(
                    "EMPTY_FILE", "The uploaded file is empty"
                ) from exc
            count = 0
            for row_number, row in enumerate(reader, start=2):
                if not row or not any(str(value).strip() for value in row):
                    continue
                count += 1
                if count > settings.max_import_rows:
                    raise IngestionError(
                        "TOO_MANY_ROWS",
                        f"Files may contain at most {settings.max_import_rows} data rows",
                        413,
                    )
                if len(row) != len(headers):
                    raise IngestionError(
                        "MALFORMED_ROW",
                        f"Row {row_number} has {len(row)} fields; expected {len(headers)}",
                    )
                for value in row:
                    _validate_cell(value)
                yield ParsedRow(
                    row_number, dict(zip(headers, row, strict=True)), frozenset()
                )
    except UnicodeDecodeError as exc:
        raise IngestionError(
            "INVALID_ENCODING", "CSV files must use UTF-8 encoding"
        ) from exc
    except csv.Error as exc:
        raise IngestionError("MALFORMED_CSV", "The CSV structure is malformed") from exc


def _validate_xlsx_archive(path: Path) -> None:
    if not zipfile.is_zipfile(path):
        raise IngestionError(
            "MALFORMED_XLSX", "The XLSX file is not a valid workbook archive"
        )
    try:
        with zipfile.ZipFile(path) as archive:
            total = 0
            for item in archive.infolist():
                normalized = item.filename.replace("\\", "/")
                if normalized.startswith("/") or "../" in normalized:
                    raise IngestionError(
                        "UNSAFE_WORKBOOK",
                        "The workbook contains an unsafe archive path",
                    )
                if normalized.lower().endswith("vbaproject.bin"):
                    raise IngestionError(
                        "MACROS_NOT_ALLOWED",
                        "Macro-enabled workbook content is not supported",
                    )
                total += item.file_size
                if total > settings.max_workbook_uncompressed_bytes:
                    raise IngestionError(
                        "WORKBOOK_TOO_LARGE",
                        "The workbook expands beyond the configured safety limit",
                        413,
                    )
    except zipfile.BadZipFile as exc:
        raise IngestionError("MALFORMED_XLSX", "The XLSX archive is malformed") from exc


def workbook_sheets(path: Path) -> list[str]:
    _validate_xlsx_archive(path)
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise IngestionError(
            "MALFORMED_XLSX", "The XLSX workbook cannot be opened"
        ) from exc
    try:
        sheets = list(workbook.sheetnames)
        if not sheets:
            raise IngestionError(
                "NO_WORKSHEETS", "The workbook does not contain any worksheets"
            )
        return sheets
    finally:
        workbook.close()


def _iter_xlsx(path: Path, sheet_name: str | None) -> Iterator[ParsedRow]:
    _validate_xlsx_archive(path)
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise IngestionError(
            "MALFORMED_XLSX", "The XLSX workbook cannot be opened"
        ) from exc
    try:
        if not workbook.sheetnames:
            raise IngestionError(
                "NO_WORKSHEETS", "The workbook does not contain any worksheets"
            )
        selected = sheet_name or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise IngestionError(
                "SHEET_NOT_FOUND", "The selected worksheet does not exist", 404
            )
        rows = workbook[selected].iter_rows()
        try:
            header_cells = next(rows)
        except StopIteration as exc:
            raise IngestionError(
                "EMPTY_SHEET", "The selected worksheet is empty"
            ) from exc
        headers = _validate_headers([cell.value for cell in header_cells])
        count = 0
        for row_number, cells in enumerate(rows, start=2):
            values = [cell.value for cell in cells[: len(headers)]]
            if len(cells) > len(headers) and any(
                cell.value is not None for cell in cells[len(headers) :]
            ):
                raise IngestionError(
                    "MALFORMED_ROW",
                    f"Row {row_number} contains values beyond the header columns",
                )
            values += [None] * (len(headers) - len(values))
            if not any(value is not None and str(value).strip() for value in values):
                continue
            count += 1
            if count > settings.max_import_rows:
                raise IngestionError(
                    "TOO_MANY_ROWS",
                    f"Files may contain at most {settings.max_import_rows} data rows",
                    413,
                )
            formulas = set()
            record: dict[str, Any] = {}
            for header, cell, value in zip(headers, cells, values, strict=False):
                _validate_cell(value)
                if cell.data_type == "f":
                    formulas.add(header)
                    record[header] = None
                else:
                    record[header] = _safe_value(value)
            yield ParsedRow(row_number, record, frozenset(formulas))
    finally:
        workbook.close()


def iter_records(
    path: Path, file_type: str, sheet_name: str | None = None
) -> Iterator[ParsedRow]:
    if file_type == "CSV":
        yield from _iter_csv(path)
    elif file_type == "XLSX":
        yield from _iter_xlsx(path, sheet_name)
    else:
        raise IngestionError(
            "UNSUPPORTED_FILE_TYPE", "Only CSV and XLSX files are supported", 415
        )


def _value_type(value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, (int, float)):
        return "NUMBER"
    if isinstance(value, (datetime, date)):
        return "DATE_TIME"
    text = str(value).strip()
    if text.lower() in {"true", "false", "yes", "no"}:
        return "BOOLEAN"
    if CURRENCY_PATTERN.match(text) and any(
        symbol in text.upper()
        for symbol in ("$", "€", "£", "¥", "₫", "VND", "USD", "EUR", "GBP")
    ):
        return "CURRENCY"
    try:
        float(text.replace(",", ""))
        return "NUMBER"
    except ValueError:
        pass
    try:
        datetime.fromisoformat(text.replace("Z", "+00:00"))
        return "DATE_TIME"
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            datetime.strptime(text[:19], fmt)
            return "DATE_TIME"
        except ValueError:
            continue
    return "STRING"


def inspect_file(
    path: Path, file_type: str, sheet_name: str | None = None
) -> dict[str, Any]:
    sheets = workbook_sheets(path) if file_type == "XLSX" else []
    selected_sheet = sheet_name or (sheets[0] if sheets else None)
    preview: list[dict[str, Any]] = []
    type_votes: dict[str, Counter[str]] = {}
    columns: list[str] = []
    row_count = 0
    formula_cells = 0
    for parsed in iter_records(path, file_type, selected_sheet):
        row_count += 1
        if not columns:
            columns = list(parsed.values)
            type_votes = {column: Counter() for column in columns}
        formula_cells += len(parsed.formula_fields)
        if len(preview) < settings.preview_rows:
            preview.append(
                {column: _safe_value(value) for column, value in parsed.values.items()}
            )
            for column, value in parsed.values.items():
                inferred = _value_type(value)
                if inferred:
                    type_votes[column][inferred] += 1
    if row_count == 0:
        raise IngestionError("NO_DATA_ROWS", "The file does not contain any data rows")
    inferred_types = {
        column: (votes.most_common(1)[0][0] if votes else "STRING")
        for column, votes in type_votes.items()
    }
    return {
        "sheet_names": sheets,
        "selected_sheet": selected_sheet,
        "columns": columns,
        "inferred_types": inferred_types,
        "rows": preview,
        "preview_row_count": len(preview),
        "row_count": row_count,
        "column_count": len(columns),
        "formula_cells_ignored": formula_cells,
    }
