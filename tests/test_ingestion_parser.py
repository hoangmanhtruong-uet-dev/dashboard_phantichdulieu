import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from backend.ingestion.parser import inspect_file, iter_records


class IngestionParserUnitTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_csv_preview_infers_columns_types_and_rows(self):
        path = self.root / "orders.csv"
        path.write_text(
            "created_at,amount,converted\n2026-08-01T10:00:00Z,12.5,true\n",
            encoding="utf-8",
        )
        preview = inspect_file(path, "CSV")
        self.assertEqual(["created_at", "amount", "converted"], preview["columns"])
        self.assertEqual("DATE_TIME", preview["inferred_types"]["created_at"])
        self.assertEqual("NUMBER", preview["inferred_types"]["amount"])
        self.assertEqual("BOOLEAN", preview["inferred_types"]["converted"])
        self.assertEqual(1, preview["row_count"])

    def test_xlsx_formula_is_never_evaluated(self):
        path = self.root / "formula.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["created_at", "amount"])
        sheet.append(["2026-08-01", "=1+1"])
        workbook.save(path)

        row = next(iter_records(path, "XLSX"))
        self.assertIsNone(row.values["amount"])
        self.assertEqual(frozenset({"amount"}), row.formula_fields)
        self.assertEqual(1, inspect_file(path, "XLSX")["formula_cells_ignored"])

    def test_xlsx_sheet_names_are_available(self):
        path = self.root / "sheets.xlsx"
        workbook = Workbook()
        workbook.active.title = "Orders"
        workbook.active.append(["created_at", "amount"])
        workbook.active.append(["2026-08-01", 10])
        workbook.create_sheet("Archive")
        workbook.save(path)

        preview = inspect_file(path, "XLSX", "Orders")
        self.assertEqual(["Orders", "Archive"], preview["sheet_names"])
        self.assertEqual("Orders", preview["selected_sheet"])


if __name__ == "__main__":
    unittest.main()
